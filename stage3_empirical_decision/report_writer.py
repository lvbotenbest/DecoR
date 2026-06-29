from __future__ import annotations

from typing import Dict, List, Tuple, Optional


def _build_maps(
    per_eval_selected: List[dict],
    per_eval_base_models: List[dict],
) -> Tuple[Dict[str, dict], Dict[Tuple[str, str], dict], List[str], List[str]]:
    ours_by_eval = {row.get('eval_name'): row for row in per_eval_selected if row.get('eval_name') is not None}

    base_by_eval_model: Dict[Tuple[str, str], dict] = {}
    eval_names = set(ours_by_eval.keys())
    model_names = set()

    for row in per_eval_base_models:
        ename = row.get('eval_name')
        mname = row.get('model_name')
        if ename is None or mname is None:
            continue
        base_by_eval_model[(ename, mname)] = row
        eval_names.add(ename)
        model_names.add(mname)

    return ours_by_eval, base_by_eval_model, sorted(eval_names), sorted(model_names)


def _write_sheet(ws, title: str, eval_names: List[str], columns: List[str], getter):
    ws.title = title
    ws.cell(row=1, column=1, value='eval_name')
    for j, col in enumerate(columns, start=2):
        ws.cell(row=1, column=j, value=col)

    for i, ename in enumerate(eval_names, start=2):
        ws.cell(row=i, column=1, value=ename)
        for j, col in enumerate(columns, start=2):
            val = getter(ename, col)
            if val is None:
                continue
            ws.cell(row=i, column=j, value=val)


def _aggregate_grouped(
    per_eval_selected: List[dict],
    per_eval_base_models: List[dict],
    eval_groups: Dict[str, List[str]],
) -> Tuple[List[dict], List[dict]]:
    eval_to_group: Dict[str, str] = {}
    for gname, enames in eval_groups.items():
        for en in enames:
            eval_to_group[str(en).strip().upper()] = gname

    ours = {}
    for row in per_eval_selected:
        ename = row.get('eval_name')
        if ename is None or ename == 'ALL':
            continue
        g = eval_to_group.get(str(ename).strip().upper())
        if g is None:
            continue
        cnt = int(row.get('count', 0) or 0)
        sum_perf = float(row.get('avg_performance', 0.0) or 0.0) * cnt
        sum_cost = float(row.get('total_cost', 0.0) or 0.0)
        agg = ours.setdefault(g, {'sum_perf': 0.0, 'sum_cost': 0.0, 'count': 0})
        agg['sum_perf'] += sum_perf
        agg['sum_cost'] += sum_cost
        agg['count'] += cnt

    base = {}
    for row in per_eval_base_models:
        ename = row.get('eval_name')
        if ename is None or ename == 'ALL':
            continue
        g = eval_to_group.get(str(ename).strip().upper())
        if g is None:
            continue
        model_name = row.get('model_name')
        if model_name is None:
            continue
        cnt = int(row.get('count', 0) or 0)
        sum_perf = float(row.get('avg_performance', 0.0) or 0.0) * cnt
        sum_cost = float(row.get('total_cost', 0.0) or 0.0)
        agg = base.setdefault((g, model_name), {'sum_perf': 0.0, 'sum_cost': 0.0, 'count': 0})
        agg['sum_perf'] += sum_perf
        agg['sum_cost'] += sum_cost
        agg['count'] += cnt

    grouped_selected = []
    for gname, agg in ours.items():
        cnt = agg['count']
        grouped_selected.append({
            'eval_name': gname,
            'count': cnt,
            'avg_performance': (agg['sum_perf'] / cnt) if cnt else 0.0,
            'total_cost': agg['sum_cost'],
            'avg_cost': (agg['sum_cost'] / cnt) if cnt else 0.0,
        })

    grouped_base = []
    for (gname, model_name), agg in base.items():
        cnt = agg['count']
        grouped_base.append({
            'eval_name': gname,
            'model_name': model_name,
            'count': cnt,
            'avg_performance': (agg['sum_perf'] / cnt) if cnt else 0.0,
            'total_cost': agg['sum_cost'],
            'avg_cost': (agg['sum_cost'] / cnt) if cnt else 0.0,
        })

    return grouped_selected, grouped_base


def write_eval_pivot_xlsx(
    output_path: str,
    per_eval_selected: List[dict],
    per_eval_base_models: List[dict],
    ours_column_name: str = 'ours',
    eval_groups: Optional[Dict[str, List[str]]] = None,
    eval_group_order: Optional[List[str]] = None,
) -> None:
    try:
        from openpyxl import Workbook
    except Exception as e:  # pragma: no cover
        raise RuntimeError(
            "openpyxl is required to write .xlsx pivot report; please install openpyxl or use --excel_output with .csv"
        ) from e

    ours_by_eval, base_by_eval_model, eval_names, model_names = _build_maps(
        per_eval_selected=per_eval_selected,
        per_eval_base_models=per_eval_base_models,
    )

    columns = list(model_names)
    if ours_column_name not in columns:
        columns.append(ours_column_name)

    wb = Workbook()

    ws_perf = wb.active

    def get_avg_perf(ename: str, col: str):
        if col == ours_column_name:
            row = ours_by_eval.get(ename)
            return None if row is None else row.get('avg_performance')
        row = base_by_eval_model.get((ename, col))
        return None if row is None else row.get('avg_performance')

    _write_sheet(ws_perf, 'avg_performance', eval_names, columns, get_avg_perf)

    ws_avg_cost = wb.create_sheet('avg_cost')

    def get_avg_cost(ename: str, col: str):
        if col == ours_column_name:
            row = ours_by_eval.get(ename)
            return None if row is None else row.get('avg_cost')
        row = base_by_eval_model.get((ename, col))
        return None if row is None else row.get('avg_cost')

    _write_sheet(ws_avg_cost, 'avg_cost', eval_names, columns, get_avg_cost)

    ws_total_cost = wb.create_sheet('total_cost')

    def get_total_cost(ename: str, col: str):
        if col == ours_column_name:
            row = ours_by_eval.get(ename)
            return None if row is None else row.get('total_cost')
        row = base_by_eval_model.get((ename, col))
        return None if row is None else row.get('total_cost')

    _write_sheet(ws_total_cost, 'total_cost', eval_names, columns, get_total_cost)

    ws_count = wb.create_sheet('count')

    def get_count(ename: str, col: str):
        if col == ours_column_name:
            row = ours_by_eval.get(ename)
            return None if row is None else row.get('count')
        row = base_by_eval_model.get((ename, col))
        return None if row is None else row.get('count')

    _write_sheet(ws_count, 'count', eval_names, columns, get_count)

    if eval_groups:
        grouped_selected, grouped_base = _aggregate_grouped(
            per_eval_selected=per_eval_selected,
            per_eval_base_models=per_eval_base_models,
            eval_groups=eval_groups,
        )
        ours_g, base_g, _eval_names_g, _model_names_g = _build_maps(
            per_eval_selected=grouped_selected,
            per_eval_base_models=grouped_base,
        )
        group_eval_names = eval_group_order if eval_group_order else sorted(list(ours_g.keys()))

        ws_g_perf = wb.create_sheet('grouped_avg_performance')

        def g_get_avg_perf(ename: str, col: str):
            if col == ours_column_name:
                row = ours_g.get(ename)
                return None if row is None else row.get('avg_performance')
            row = base_g.get((ename, col))
            return None if row is None else row.get('avg_performance')

        _write_sheet(ws_g_perf, 'grouped_avg_performance', group_eval_names, columns, g_get_avg_perf)

        ws_g_avg_cost = wb.create_sheet('grouped_avg_cost')

        def g_get_avg_cost(ename: str, col: str):
            if col == ours_column_name:
                row = ours_g.get(ename)
                return None if row is None else row.get('avg_cost')
            row = base_g.get((ename, col))
            return None if row is None else row.get('avg_cost')

        _write_sheet(ws_g_avg_cost, 'grouped_avg_cost', group_eval_names, columns, g_get_avg_cost)

        ws_g_total_cost = wb.create_sheet('grouped_total_cost')

        def g_get_total_cost(ename: str, col: str):
            if col == ours_column_name:
                row = ours_g.get(ename)
                return None if row is None else row.get('total_cost')
            row = base_g.get((ename, col))
            return None if row is None else row.get('total_cost')

        _write_sheet(ws_g_total_cost, 'grouped_total_cost', group_eval_names, columns, g_get_total_cost)

        ws_g_count = wb.create_sheet('grouped_count')

        def g_get_count(ename: str, col: str):
            if col == ours_column_name:
                row = ours_g.get(ename)
                return None if row is None else row.get('count')
            row = base_g.get((ename, col))
            return None if row is None else row.get('count')

        _write_sheet(ws_g_count, 'grouped_count', group_eval_names, columns, g_get_count)

    wb.save(output_path)
